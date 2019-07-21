# -*-coding:utf-8-*-
"""
****************************************
author:善待今天
time:2019/07/01 23:10
file:read_write_excel.py
software:PyCharm Community Edition
E-mail:2904504961@qq.com
Motivational motto: Do difficult things and get something
****************************************
"""

import openpyxl


class Case:
    def __init__(self,attrs):
        for i in attrs:
            setattr(self,i[0],i[1])


class ReadExcel(object):
    """
    封装一个读取Excel表的类
    """
    def __init__(self,file_name,sheet_name):
        """
        用来初始化对象
        :param file_name:   读取的文件名，str类型
        :param sheet_name:  读取的表单名，str类型
        """
        self.file_name = file_name
        self.wb = openpyxl.load_workbook(file_name)
        self.sh = self.wb[sheet_name]

    def __del__(self):
        """
        在对象销毁之后执行
        :return:
        """
        # 关闭文件
        self.wb.close()

    def r_excel(self):
        """
        用来读取Excel中的数据
        :return:cases,每条测试用例以字典的形式存在列表中
        """
        # 按行获取数据转换成列表
        rows_data = list(self.sh.rows)

        # 获取表头
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)

        # 定义空列表cases = []用来储存测试用例数据
        cases = []
        for case in rows_data[1:]:
            # 定义空列表case_data = []用来储存每条测试用例数据
            case_data = []
            for cell in case:
                case_data.append(cell.value)
            cases.append(dict(zip(titles, case_data)))
        return cases

    def r_excel_obj(self):
        """
        按行读取Excel中的数据
        :return: cases，每条测试用例以对象的形式存在列表中
        """
        # 按行读取数据，转换成列表
        rows_data = list(self.sh.rows)

        # 获取表头
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)

        cases = []  # 用来存放测试用例
        for data in rows_data[1:]:
            case_data = []  # 存放每条测试用例数据
            for cell in data:
                case_data.append(cell.value)
            case = list(zip(titles, case_data))
            case_obj = Case(case)
            cases.append(case_obj)
        return cases

    def read_excel(self,list0):
        """
        读取指定列的数据，并且以字典的形式存在列表中
        :param list0:   传入的指定列，list类型
        :return:    cases，list类型，每个元素为一条测试用例
        """
        list0 = eval(list0)
        if len(list0) == 0:
            return self.r_excel_obj()
        rows = self.sh.max_row       # 获取行总数rows
        titles = []                     # 空列表titles用来存放表头
        cases = []                      # 空列表cases用来存放所有用例
        for row in range(1,rows+1):
            if row == 1:
                for column in list0:
                    title = self.sh.cell(row,column).value
                    titles.append(title)
            else:
                case_data = []  # 空列表case_data用来存放单条测试用例
                for column in list0:
                    cell_data = self.sh.cell(row,column).value
                    case_data.append(cell_data)
                case = dict(zip(titles,case_data))
                cases.append(case)
        return cases

    def read_excel_obj(self,list0):
        """
        读取指定列的数据，并且以对象的形式存在列表中
        :param list0: 读取的列，list类型
        :return:
        """
        list0 = eval(list0)
        if len(list0) == 0 or [1,2,3] not in list0:
            return self.r_excel_obj()
        rows = self.sh.max_row              # 获取行总数rows
        titles = []                         # 空列表titles存放表头
        cases = []                          # 空列表cases用来存放所有的测试用例
        for row in range(1,rows+1):
            if row == 1:
                for column in list0:
                    title = self.sh.cell(row,column).value
                    titles.append(title)
            else:
                case_data = []              # 空列表case_data用来存放单条测试用例数据
                for column in list0:
                    cell_data = self.sh.cell(row,column).value
                    case_data.append(cell_data)
                case = zip(titles,case_data)
                case_obj = Case(case)
                cases.append(case_obj)
        return cases


    def read_write_data(self, row, column, msg):
        """
        写入数据
        :param row:         写入数据的行
        :param column:      写入数据的列
        :param msg:         写入单元格的信息
        :return:
        """
        self.sh.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)


